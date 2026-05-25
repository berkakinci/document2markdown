"""XLSX converter — produces a markdown index with CSV data and images.

Each sheet becomes a CSV file stored as an EmbeddedAsset. Embedded images
are extracted per-sheet. Charts are noted but cannot be rendered (openpyxl
exposes chart XML definitions, not rasterized images).

Output structure (rendered by the standard MarkdownRenderer):
- H1: source filename (stem)
- Per sheet: H2 heading, link to CSV, inline images
- Non-exported content section (charts as UnsupportedBlock)
"""

from __future__ import annotations

import csv
import io
import logging
from pathlib import Path

from document2markdown.converter_base import BaseConverter
from document2markdown.document_model import (
    ConversionResult,
    EmbeddedAsset,
    HeadingBlock,
    ImageBlock,
    LinkBlock,
    ParagraphBlock,
    UnsupportedBlock,
)

try:
    import openpyxl
except ImportError as _err:  # pragma: no cover
    raise ImportError(
        "openpyxl is required for XLSXConverter. Install it with: pip install openpyxl"
    ) from _err

logger = logging.getLogger(__name__)


class XLSXConverter(BaseConverter):
    """Convert an ``.xlsx`` file to a :class:`ConversionResult`.

    Produces:
    - One CSV EmbeddedAsset per sheet (tabular data)
    - Image EmbeddedAssets for any embedded images per sheet
    - Markdown blocks forming an index/overview document
    """

    def convert(self, source_path: Path) -> ConversionResult:
        """Convert *source_path* (.xlsx) to a :class:`ConversionResult`."""
        blocks: list = []
        embedded: list[EmbeddedAsset] = []
        warnings: list[str] = []
        non_exported: list[str] = []

        try:
            wb = openpyxl.load_workbook(str(source_path), data_only=True)
        except Exception as exc:
            msg = f"failed to open XLSX '{source_path}': {exc}"
            logger.error(msg)
            warnings.append(msg)
            return ConversionResult(blocks=[], embedded=[], warnings=warnings)

        # Document title
        blocks.append(HeadingBlock(level=1, text=source_path.stem))

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # --- Sheet heading ---
            blocks.append(HeadingBlock(level=2, text=sheet_name))

            # --- CSV export ---
            csv_data = self._sheet_to_csv(ws)
            if csv_data:
                row_count = ws.max_row or 0
                col_count = ws.max_column or 0
                csv_asset = EmbeddedAsset(
                    data=csv_data.encode("utf-8"),
                    extension=".csv",
                    original_name=f"{source_path.stem} - {sheet_name}.csv",
                    alt_text=f"{sheet_name} ({row_count} rows × {col_count} cols)",
                    source_vector_format=None,
                )
                asset_idx = len(embedded)
                embedded.append(csv_asset)
                blocks.append(LinkBlock(
                    text=f"{sheet_name} ({row_count} rows × {col_count} cols)",
                    asset_index=asset_idx,
                ))

            # --- Embedded images ---
            for img in getattr(ws, "_images", []):
                try:
                    img_data = img._data()
                    ext = ".png"
                    if hasattr(img, "format"):
                        ext = f".{img.format.lower()}" if img.format else ".png"
                    img_asset = EmbeddedAsset(
                        data=img_data,
                        extension=ext,
                        original_name=None,
                        alt_text=f"{sheet_name} image",
                        source_vector_format=None,
                    )
                    img_idx = len(embedded)
                    embedded.append(img_asset)
                    blocks.append(ImageBlock(asset_index=img_idx, alt=img_asset.alt_text))
                except Exception as exc:
                    warnings.append(f"Failed to extract image from '{sheet_name}': {exc}")

            # --- Charts (non-exportable) ---
            charts = getattr(ws, "_charts", [])
            if charts:
                non_exported.append(
                    f"{sheet_name}: {len(charts)} chart(s) (not exportable)"
                )

        # --- Non-exported content section ---
        if non_exported:
            blocks.append(HeadingBlock(level=2, text="Non-exported content"))
            for note in non_exported:
                blocks.append(UnsupportedBlock(description=note))

        wb.close()
        return ConversionResult(blocks=blocks, embedded=embedded, warnings=warnings)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _sheet_to_csv(ws) -> str:
        """Serialize a worksheet to a CSV string. Returns empty string if sheet is empty."""
        buf = io.StringIO()
        writer = csv.writer(buf)
        has_data = False
        for row in ws.iter_rows(values_only=True):
            # Skip fully empty rows
            if any(cell is not None for cell in row):
                writer.writerow([cell if cell is not None else "" for cell in row])
                has_data = True
        return buf.getvalue() if has_data else ""
